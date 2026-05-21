"""Payroll exports — CSV, XLSX, Sage Paie CSV, Monthly PDF summary."""
from __future__ import annotations
import csv
import io
from datetime import datetime, date, timedelta, time
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response

from sqlalchemy import func, and_, cast, Date

from app.database.connection import get_db_session
from app.database.schema import (
    Attendance as DBAttendance, Employee as DBEmployee,
    Department as DBDepartment, AppSettings,
)
from app.core.security import get_current_user, require_any_permission, MANAGER_PERMS

# Payroll exports are a management feature — not for plain reporting users.
_require_manager = require_any_permission(*MANAGER_PERMS)
from app.services.punch_classifier import get_employee_day_summary

router = APIRouter()


def _parse_date(s: str) -> date:
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(400, "Invalid date format, use YYYY-MM-DD")


def _gather_rows(start: date, end: date, employee_ids: Optional[list[int]] = None):
    """Return list of dicts: {employee_id, matricule, name, department, date,
                              entry, exit, worked_min, overtime_min, late_min,
                              early_min, swipes}.
    Voided rows excluded. Respects shared/separate employee mode for grouping by user_id.
    """
    rows = []
    start_dt = datetime.combine(start, time.min)
    end_dt = datetime.combine(end, time.max)

    with get_db_session() as db:
        settings = db.query(AppSettings).first()
        shared = (getattr(settings, 'employee_mode', None) or 'shared') == 'shared'

        id_col = DBEmployee.user_id if shared else DBEmployee.id

        q = (db.query(
            id_col.label("employee_id"),
            (func.min(DBEmployee.name) if shared else DBEmployee.name).label("employee_name"),
            (func.min(DBEmployee.user_id) if shared else DBEmployee.user_id).label("matricule"),
            (func.coalesce(func.min(DBDepartment.name), "-") if shared else func.coalesce(DBDepartment.name, "-")).label("department"),
            cast(DBAttendance.timestamp, Date).label("day"),
            func.min(DBAttendance.timestamp).label("first_ts"),
            func.max(DBAttendance.timestamp).label("last_ts"),
            func.count(DBAttendance.id).label("swipes"),
        )
        .select_from(DBAttendance)
        .join(DBEmployee, DBAttendance.employee_id == DBEmployee.id)
        .outerjoin(DBDepartment, DBEmployee.department_id == DBDepartment.id)
        .filter(DBAttendance.voided_by_correction_id.is_(None))
        .filter(DBAttendance.approved.isnot(False))
        .filter(DBAttendance.timestamp >= start_dt)
        .filter(DBAttendance.timestamp <= end_dt))

        if employee_ids:
            q = q.filter(DBEmployee.id.in_(employee_ids))

        if shared:
            q = q.group_by(DBEmployee.user_id, cast(DBAttendance.timestamp, Date))
        else:
            q = q.group_by(DBEmployee.id, DBEmployee.name, DBEmployee.user_id, DBDepartment.name, cast(DBAttendance.timestamp, Date))

        agg_rows = q.order_by(cast(DBAttendance.timestamp, Date).asc()).all()

        # uid -> [pks] map for shared mode summary lookups
        uid_to_pks = {}
        if shared:
            for r in agg_rows:
                if r.employee_id not in uid_to_pks:
                    uid_to_pks[r.employee_id] = [pk for (pk,) in
                        db.query(DBEmployee.id).filter(DBEmployee.user_id == r.employee_id).all()]

        for r in agg_rows:
            day_d = r.day
            pks = uid_to_pks.get(r.employee_id, [r.employee_id]) if shared else [r.employee_id]
            summary = {}
            if pks:
                summary = get_employee_day_summary(db, pks[0], day_d, employee_ids=pks if shared else None) or {}

            entry = r.first_ts
            exit_ = r.last_ts
            if int(r.swipes or 0) == 1 and entry:
                if entry.hour < 12:
                    exit_ = None
                else:
                    entry = None

            rows.append({
                "employee_id": r.employee_id,
                "matricule": r.matricule,
                "name": r.employee_name,
                "department": r.department,
                "date": day_d.isoformat() if hasattr(day_d, 'isoformat') else str(day_d),
                "entry": entry.strftime("%H:%M") if entry else "-",
                "exit":  exit_.strftime("%H:%M")  if exit_  else "-",
                "worked_min":   int(summary.get("total_minutes") or 0),
                "overtime_min": int(summary.get("overtime_minutes") or 0),
                "late_min":     int(summary.get("late_minutes") or 0),
                "early_min":    int(summary.get("early_departure_minutes") or 0),
                "swipes": int(r.swipes or 0),
            })

    return rows


def _fmt_hm(m: int) -> str:
    if not m:
        return "0:00"
    h, mm = divmod(int(m), 60)
    return f"{h}:{mm:02d}"


# ── CSV ───────────────────────────────────────────────────────────────────────
@router.get("/payroll-export/csv")
def export_csv(start_date: str, end_date: str, current=Depends(_require_manager)):
    rows = _gather_rows(_parse_date(start_date), _parse_date(end_date))
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=';')
    w.writerow(["Matricule", "Nom", "Departement", "Date", "Entree", "Sortie",
                "Travail", "Heures sup.", "Retard", "Depart anticipe", "Pointages"])
    for r in rows:
        w.writerow([r["matricule"], r["name"], r["department"], r["date"],
                    r["entry"], r["exit"],
                    _fmt_hm(r["worked_min"]), _fmt_hm(r["overtime_min"]),
                    _fmt_hm(r["late_min"]), _fmt_hm(r["early_min"]), r["swipes"]])
    data = buf.getvalue().encode("utf-8-sig")  # BOM for Excel
    return Response(content=data, media_type="text/csv",
                    headers={"Content-Disposition": f'attachment; filename="payroll_{start_date}_{end_date}.csv"'})


# ── XLSX ──────────────────────────────────────────────────────────────────────
@router.get("/payroll-export/xlsx")
def export_xlsx(start_date: str, end_date: str, current=Depends(_require_manager)):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl not installed — run pip install openpyxl")

    rows = _gather_rows(_parse_date(start_date), _parse_date(end_date))

    wb = Workbook()
    ws = wb.active
    ws.title = "Paie"

    headers = ["Matricule", "Nom", "Département", "Date", "Entrée", "Sortie",
               "Travail", "Heures sup.", "Retard", "Départ ant.", "Pointages"]
    bold = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1E3A5F")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(headers)
    for c in ws[1]:
        c.font = bold; c.fill = head_fill; c.alignment = Alignment(horizontal="center"); c.border = border

    totals = {"worked": 0, "overtime": 0, "late": 0, "early": 0}
    for r in rows:
        ws.append([r["matricule"], r["name"], r["department"], r["date"],
                   r["entry"], r["exit"],
                   _fmt_hm(r["worked_min"]), _fmt_hm(r["overtime_min"]),
                   _fmt_hm(r["late_min"]), _fmt_hm(r["early_min"]), r["swipes"]])
        totals["worked"] += r["worked_min"]
        totals["overtime"] += r["overtime_min"]
        totals["late"] += r["late_min"]
        totals["early"] += r["early_min"]

    # Totals row
    last = ws.max_row + 1
    ws.cell(row=last, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=last, column=7, value=_fmt_hm(totals["worked"])).font = Font(bold=True)
    ws.cell(row=last, column=8, value=_fmt_hm(totals["overtime"])).font = Font(bold=True)
    ws.cell(row=last, column=9, value=_fmt_hm(totals["late"])).font = Font(bold=True)
    ws.cell(row=last, column=10, value=_fmt_hm(totals["early"])).font = Font(bold=True)

    # Column widths
    widths = [12, 28, 18, 12, 9, 9, 10, 12, 9, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    buf = io.BytesIO(); wb.save(buf)
    return Response(content=buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="payroll_{start_date}_{end_date}.xlsx"'})


# ── Sage Paie Maroc ───────────────────────────────────────────────────────────
@router.get("/payroll-export/sage-paie")
def export_sage(start_date: str, end_date: str, current=Depends(_require_manager)):
    """Best-effort Sage Paie Maroc-compatible CSV.
    Layout (per row): MATRICULE;CODE_RUBRIQUE;NB_HEURES;DATE_DEBUT;DATE_FIN
    Rubriques used (configurable in Sage — adjust on import if needed):
      HEUR — heures normales
      HSUP — heures supplémentaires
      RTRD — minutes de retard
    """
    rows = _gather_rows(_parse_date(start_date), _parse_date(end_date))
    # Aggregate per employee
    agg: dict = {}
    for r in rows:
        k = r["matricule"] or "-"
        a = agg.setdefault(k, {"worked": 0, "overtime": 0, "late": 0})
        a["worked"] += r["worked_min"]
        a["overtime"] += r["overtime_min"]
        a["late"] += r["late_min"]

    buf = io.StringIO()
    w = csv.writer(buf, delimiter=';')
    w.writerow(["MATRICULE", "RUBRIQUE", "QUANTITE", "DATE_DEBUT", "DATE_FIN"])
    for matricule, a in sorted(agg.items()):
        if a["worked"]:
            w.writerow([matricule, "HEUR", f"{a['worked']/60:.2f}", start_date, end_date])
        if a["overtime"]:
            w.writerow([matricule, "HSUP", f"{a['overtime']/60:.2f}", start_date, end_date])
        if a["late"]:
            w.writerow([matricule, "RTRD", str(a["late"]), start_date, end_date])
    data = buf.getvalue().encode("utf-8-sig")
    return Response(content=data, media_type="text/csv",
                    headers={"Content-Disposition": f'attachment; filename="sage_paie_{start_date}_{end_date}.csv"'})


# ── Monthly PDF summary ───────────────────────────────────────────────────────
@router.get("/payroll-export/monthly-pdf")
def export_monthly_pdf(year: int, month: int, current=Depends(_require_manager)):
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors

    if not (1 <= month <= 12):
        raise HTTPException(400, "Invalid month")
    start_d = date(year, month, 1)
    end_d = (date(year + (month == 12), (month % 12) + 1, 1) - timedelta(days=1))

    rows = _gather_rows(start_d, end_d)

    # Group per employee
    per_emp: dict = {}
    for r in rows:
        e = per_emp.setdefault(r["employee_id"],
            {"name": r["name"], "matricule": r["matricule"], "department": r["department"],
             "days": 0, "worked": 0, "overtime": 0, "late": 0, "early": 0})
        e["days"] += 1
        e["worked"] += r["worked_min"]
        e["overtime"] += r["overtime_min"]
        e["late"] += r["late_min"]
        e["early"] += r["early_min"]

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=24, rightMargin=24, topMargin=28, bottomMargin=24)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('t', parent=styles['Heading1'], fontSize=16, alignment=1)
    elements = []

    elements.append(Paragraph(f"Récapitulatif mensuel — {month:02d}/{year}", title_style))
    elements.append(Spacer(1, 12))

    head = ["Matricule", "Nom", "Département", "Jours", "Travail", "H. sup.", "Retard", "Dép. ant."]
    data = [head]
    totals = {"days": 0, "worked": 0, "overtime": 0, "late": 0, "early": 0}
    for e in sorted(per_emp.values(), key=lambda x: x["name"] or ""):
        data.append([e["matricule"] or "-", e["name"] or "-", e["department"] or "-",
                     e["days"], _fmt_hm(e["worked"]), _fmt_hm(e["overtime"]),
                     _fmt_hm(e["late"]), _fmt_hm(e["early"])])
        for k in totals: totals[k] += e[k] if k != "days" else e["days"]

    data.append(["TOTAL", "", "", totals["days"], _fmt_hm(totals["worked"]),
                 _fmt_hm(totals["overtime"]), _fmt_hm(totals["late"]), _fmt_hm(totals["early"])])

    tbl = Table(data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 9),
        ('GRID',       (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F3F4F6')),
        ('FONTNAME',   (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#FAFAFA')]),
    ]))
    elements.append(tbl)

    doc.build(elements)
    return Response(content=buf.getvalue(), media_type="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="monthly_summary_{year}_{month:02d}.pdf"'})
